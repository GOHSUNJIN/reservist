#!/usr/bin/env python3
"""Generate testing_checklist.xlsx from structured test data."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SEC_BG = "2c3e50"; SEC_FG = "FFFFFF"
HDR_BG = "2f5fd0"; HDR_FG = "FFFFFF"; HDR_BR = "1e3fa0"
ROW_O  = "FFFFFF"; ROW_E  = "f5f7fb"; ROW_BR = "e0e4ec"
TXT    = "1a2233"; TXT_ID = "9aa3b2"; TXT_NT = "5c6678"
PH_BG  = "fdecea"; PH_FG  = "c0392b"
PM_BG  = "fef6e4"; PM_FG  = "b9791a"
PL_BG  = "eaf5f0"; PL_FG  = "1f8a5b"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(color="000000", sz=10, bold=False):
    return Font(name="Arial", size=sz, bold=bold, color=color)
def _al(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _br(color=ROW_BR):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, bottom=s)
def _sec_br():
    return Border(bottom=Side(style="thin", color="1a2233"))

SECTIONS = [
    ("1. AUTHENTICATION AND SESSION", [
        ("1.1",  "Login",                         "Log in with a valid contact number and password",
                 "User is authenticated and taken to the correct tab (Overview for admin, Check-in for reservist)", "High"),
        ("1.2",  "Login - wrong password",         "Enter a correct contact but wrong password",
                 'Error: "Invalid contact number or password." No login occurs', "High"),
        ("1.3",  "Login - blank fields",           "Attempt login with empty contact and/or password",
                 'Error: "Enter your contact number." No request sent', "Med"),
        ("1.4",  "Logout",                         "Tap logout while authenticated",
                 "Confirmation dialog appears. On confirm: session cleared, returned to login screen, no state persists", "High"),
        ("1.5",  "Idle timeout",                   "Leave the app idle for 20 minutes without interaction",
                 "Session expires, toast shown, returned to login screen automatically", "High"),
        ("1.6",  "Idle warning (18 min)",           "Leave the app idle for 18 minutes without interaction",
                 'Warning banner appears: idle warning shown with option to stay. Session has not yet ended', "Med"),
        ("1.7",  "Session expiry warning",         "Wait 55 minutes after login without refreshing the session",
                 'Banner shown: "Your session is about to expire." Tap to stay logged in resets the timer', "Med"),
        ("1.8",  "Tab persistence",                "Close and reopen the browser tab before the session expires",
                 "User is still logged in. Session state is restored from sessionStorage. No re-login required", "Med"),
        ("1.9",  "Self-signup - new user",         "Sign up with a valid name, Singapore number (starts 6/8/9, 8 digits), and password (min 6 chars)",
                 'Screen shows "Your request has been submitted." Auth account is created. Pending signup appears in admin inbox labelled New', "High"),
        ("1.10", "Signup - invalid contact",       "Attempt signup with a contact that does not start with 6, 8, or 9, or has wrong digit count",
                 'Error: "Contact must be an 8-digit Singapore number." No account created', "High"),
        ("1.11", "Signup - short password",        "Attempt signup with a password under 6 characters",
                 'Error: "Password must be at least 6 characters." No account created', "Med"),
        ("1.12", "Signup - already pending",       "Attempt signup again for a contact that already has a pending signup request",
                 'Error: "A signup request for this number is already pending admin approval."', "High"),
        ("1.13", "Signup - rejected contact",      "Attempt signup with a contact that was previously rejected",
                 'Error: "Your previous signup request was not approved. Contact your supervisor."', "High"),
        ("1.14", "Signup - outside active cycle",  "Attempt signup when no cycle is currently live",
                 'Error: "No active intake batch is open for sign-up right now."', "Med"),
        ("1.15", "Login - inactive account",       "Log in with a valid contact that belongs to an inactive (deactivated) reservist",
                 'Re-enrollment request auto-submitted. Error shown: "Your account is inactive... A re-enrollment request has been sent."', "High"),
        ("1.16", "Demo - reservist mode",          'Tap "Try as Reservist" on the login screen',
                 "Logged in as Demo User with 4 demo personnel. Check-in tab is active. No DB calls are made", "Low"),
        ("1.17", "Demo - admin mode",              'Tap "Try as Supervisor" on the login screen',
                 "Logged in as Supervisor with demo attendance data. Overview tab is active. No DB calls are made", "Low"),
    ]),
    ("2. CHECK-IN - OPS SECURITY (4 PHASES)", [
        ("2.1",  "GPS verification",               "Tap Verify Location while physically within the HQ radius",
                 'Status changes to "Verified" with green indicator. Distance from HQ is shown', "High"),
        ("2.2",  "GPS out of range",               "Tap Verify Location while outside the HQ radius",
                 'Status shows "Out of range" with distance displayed. Check-in button is not available', "High"),
        ("2.3",  "Phase 1 check-in",               "With GPS verified, tap the Check In button for Phase 1",
                 'Toast: "Checked in." Status card updates. Time and distance are recorded. Phase 2 tile becomes available', "High"),
        ("2.4",  "Phase outside time window",      "View the check-in screen when the current time is outside all phase windows",
                 "Phase tile is visible but the action button is greyed out and cannot be tapped. No error occurs", "Med"),
        ("2.5",  "Phase prerequisite enforcement", "Attempt Phase 3 (Return) before Phase 2 (Lunch out) has been logged",
                 'Error toast: "Record your break out first." Phase 3 is not submitted', "High"),
        ("2.6",  "All 4 phases complete",          "Submit all four phases in order within a single day",
                 'Completion card shown. All phase times are visible. Status is "Present". Meal eligibility updates if applicable', "High"),
        ("2.7",  "GPS bypass trigger",             "Tap Verify Location and fail GPS 5 consecutive times",
                 'After the 5th failure, a "No GPS" bypass option appears. Before 5 failures, the bypass is not shown', "Med"),
        ("2.8",  "GPS bypass check-in",            'Use the "No GPS" bypass option to check in without GPS verification',
                 "Check-in recorded. Attendance record is flagged as GPS bypassed in the log. Distance is null", "Med"),
        ("2.9",  "Late check-in warning (30 min)", "Check in between 09:30 and 10:00 (30-60 min late)",
                 'Warning banner appears: "You are checking in late." Banner can be dismissed', "Med"),
        ("2.10", "Late check-in reason (1 hour)",  "Check in after 10:00 (more than 60 min late)",
                 "Reason modal opens automatically. Reason is required. Saved reason is visible in the admin log alongside a late flag", "High"),
        ("2.11", "Check-in before cycle start",    "Attempt to check in when today is before the reservist's batch start date",
                 'Error: "Your reporting period has not started yet." No record is created', "High"),
        ("2.12", "Check-in after cycle end",       "Attempt to check in when today is after the reservist's batch end date",
                 'Error: "Your reporting period has ended." No record is created', "High"),
        ("2.13", "Offline check-in",               "Enable airplane mode, then check in with GPS verified",
                 'Toast: "Checked in." Offline badge appears. On reconnection, the queued action is synced automatically and the badge disappears', "High"),
        ("2.14", "Offline queue retry limit",      "Simulate 3 failed sync attempts for a queued item (requires mocking a DB error)",
                 "After 3 failures the item is dropped from the queue. User is notified via error toast on each failed attempt", "Med"),
        ("2.15", "Deactivated account check-in",   "Attempt to check in with a user whose is_active flag is false",
                 'Error: "Your account has been deactivated. Please contact your supervisor."', "High"),
        ("2.16", "No-report day - check-in screen","Log in on a day that has been marked as a no-reporting day",
                 'Check-in status shows "No reporting" with correct label. Phase tiles are hidden. No check-in is possible', "Med"),
        ("2.17", "Weekend check-in screen",        "Open the app on a Saturday or Sunday",
                 'Status shows "Weekend - No reporting required." No phases displayed', "Med"),
    ]),
    ("3. CHECK-IN - CAS / CRIME ALERT (2 PHASES)", [
        ("3.1",  "CAS - Phase 1 check-in",         "Log in as a CAS reservist and tap Verify Location, then Check In for Phase 1",
                 "GPS verified. Check-in recorded with timestamp and distance. Phase 2 tile becomes available", "High"),
        ("3.2",  "CAS - Phase 2 check-out",        "After Phase 1 is recorded, tap Check Out for Phase 2",
                 'Phase 2 timestamp recorded. Completion card shown. Status is "Present"', "High"),
        ("3.3",  "CAS - phases 3 and 4 absent",    "Log in as a CAS reservist and inspect the check-in screen",
                 "Only 2 phase tiles are visible (Check In and Check Out). Lunch out and Return tiles are not shown", "High"),
    ]),
    ("4. LEAVE AND MC REQUESTS", [
        ("4.1",  "Submit MC request",              "Submit an MC request for a future weekday within the current cycle",
                 'Toast: "Request submitted for approval." MC badge shown in amber. Admin inbox shows new request with amber badge', "High"),
        ("4.2",  "Submit Personal Leave request",  "Submit a Personal Leave request for a future weekday within the current cycle",
                 'Toast: "Request submitted for approval." Personal Leave badge shown in red. Admin inbox shows new request with red badge', "High"),
        ("4.3",  "Submit request - already pending","Attempt to submit a second leave request while one is already pending",
                 'Error: "You already have a pending request." No second request is created', "High"),
        ("4.4",  "Submit request - past date",     "Attempt to submit a leave request for a date before today",
                 'Error: "Cannot submit a request for a past date." No request is created', "High"),
        ("4.5",  "Submit request - weekend",       "Attempt to submit a leave request for a Saturday or Sunday",
                 'Error shown: weekend dates are not valid for leave requests. No request created', "Med"),
        ("4.6",  "Submit request - public holiday","Attempt to submit a leave request for a Singapore public holiday",
                 'Error shown: public holiday dates are not valid. No request created', "Med"),
        ("4.7",  "Submit request - outside cycle", "Attempt to submit a leave request for a date outside the reservist's batch start and end dates",
                 'Error: "The selected date is outside your current cycle."', "High"),
        ("4.8",  "Submit request - duplicate date","Submit a request for a date that already has a non-cancelled, non-rejected request",
                 'Error: "You already submitted a request for this date." No duplicate is created', "High"),
        ("4.9",  "Approved leave blocks check-in", "Have an admin approve a leave request for today, then attempt to check in as that reservist",
                 "Check-in screen shows that leave is approved for today. Phase tiles are hidden or action is blocked", "High"),
        ("4.10", "Cancel pending request",         "Cancel a pending leave request before the admin has reviewed it",
                 'Toast: "Request cancelled." Status in history changes to Cancelled. DB record status is updated. Pending request badge clears', "High"),
        ("4.11", "Cancel already-reviewed request","Attempt to cancel a request that an admin has already approved or rejected",
                 'Error: "Request was already reviewed by an admin." History reloads to show the current status', "High"),
        ("4.12", "Admin approve MC leave",         "Admin approves an MC leave request from the Requests inbox",
                 "Request moves to Approved. Attendance record for that date is updated to MC. If the date is today, the roster updates immediately", "High"),
        ("4.13", "Admin reject leave with reason", "Admin rejects a leave request and provides a written reason",
                 "Request moves to Rejected. Rejection reason is recorded. Reservist sees the rejection reason in their history", "High"),
        ("4.14", "Admin bulk approve leaves",      "Select 3 pending leave requests and bulk-approve them",
                 'Toast: "3 requests approved." All 3 move to Approved. Attendance records updated for each', "Med"),
        ("4.15", "Admin bulk reject leaves",       "Select multiple pending leave requests and bulk-reject with a reason",
                 "All selected requests rejected. Each record stores the same rejection reason. Toast confirms count", "Med"),
        ("4.16", "Void leave on mark present",     "Admin marks a reservist with an approved leave as Present via the roster or log",
                 "Approved leave is voided in DB. approvedLeavesCache for that person/date is cleared. Reservist can submit a new request for that date", "High"),
        ("4.17", "Re-submit after void",           "After admin has voided a reservist's approved leave by marking them present, reservist submits a new request for the same date",
                 "New request created without duplicate error. Appears in admin inbox as pending", "High"),
        ("4.18", "Request deduplication in history","Reservist has an approved request and a later withdrawn request for the same date",
                 "Only one entry is shown in the request history for that date. The approved entry takes priority over the withdrawn one", "Med"),
        ("4.19", "History type badges",            "View the request history in the Info tab after submitting both an MC and a Personal Leave request",
                 'MC entries show an amber badge. Personal Leave entries show a red badge. Status badges (Submitted, Approved, Declined, Withdrawn) are correct', "Med"),
        ("4.20", "Declined request shows reason",  "Admin rejects a request with a written reason; reservist then views their request history",
                 "The declined request row shows the supervisor's written reason inline in the history view", "Med"),
    ]),
    ("5. SIGNUP APPROVAL WORKFLOW", [
        ("5.1",  "Approve new signup",             "Admin approves a pending signup for a brand-new user",
                 'Personnel record created and linked to the auth account. Person appears on the roster. Toast: "[Name] approved and added to the roster."', "High"),
        ("5.2",  "Approve returning reservist",    "Admin approves a signup where the contact matches an inactive personnel record (returning reservist)",
                 "Existing record is reactivated and assigned to the new cycle. Auth is linked. Person appears on roster labelled Returning", "High"),
        ("5.3",  "Approve - already on roster",    "Admin approves a signup where the contact already has an active personnel record",
                 "Auth account is linked to the existing record. No duplicate personnel entry is created", "Med"),
        ("5.4",  "Approve - personnel.add fails",  "Simulate a DB failure on personnel.add after the signup has been marked approved (requires mocking)",
                 "Signup status is reverted to Pending. Error toast shown. Admin can retry. No orphaned approved record", "High"),
        ("5.5",  "Reject signup",                  "Admin rejects a pending signup request",
                 "Request moves to Rejected tab. Auth account is retained (so the person can re-apply). No personnel record is created", "High"),
        ("5.6",  "Reopen rejected signup",         "Admin re-opens a previously rejected signup",
                 "Signup moves back to Pending tab. Admin can approve or reject again", "Med"),
        ("5.7",  "Bulk approve signups",           "Select 3 pending signups and bulk-approve them",
                 'All 3 approved. Each person added to the roster. Toast: "3 signups approved." Personnel list refreshes', "Med"),
        ("5.8",  "Signup search and filter",       "Use the collapsible search bar and filter pills (All, New, Returning) in the signups panel",
                 "Results update correctly. Red dot appears on the search icon when a filter is active. Clear resets all filters", "Med"),
        ("5.9",  "Select-all checkbox",            "Tick the select-all checkbox in the signups header with search active",
                 "Only visible (filtered) signups are selected. Unchecking deselects all. Count badge updates correctly", "Med"),
    ]),
    ("6. ADMIN - ROSTER AND ATTENDANCE MANAGEMENT", [
        ("6.1",  "Manual status override - present","Admin manually marks a reservist as Present on today's roster",
                 'Status updates immediately. DB record is upserted. Toast: "Marked present"', "High"),
        ("6.2",  "Manual status override - past",  "Navigate to a past date (viewOffset negative) and manually override a status",
                 "Status updates in attendanceCache for that date. DB record updated. Does not affect today's attendance", "Med"),
        ("6.3",  "Mark all absent",                'Use "Mark All Absent" on a day where some reservists have no status',
                 'Only reservists with no status or "pending" status are marked absent. Those already marked Present, MC, or Absent are unaffected', "High"),
        ("6.4",  "Mark all absent - future blocked","Navigate to a future date (viewOffset positive) and attempt \"Mark All Absent\"",
                 "Action is blocked. No status changes. Confirm dialog closes immediately", "High"),
        ("6.5",  "Mark all present",               'Use "Mark All Present" on today\'s roster',
                 "All reservists with no status are marked Present with current time as check-in. Those already marked are unaffected", "Med"),
        ("6.6",  "Manual time correction",         "Edit a reservist's times for all 4 phases from the Log tab",
                 "Times saved. Record flagged as GPS bypassed. Edit log entry created with editor name and timestamp", "High"),
        ("6.7",  "Time correction - invalid order","Enter a Phase 2 time that is earlier than Phase 1",
                 'Error toast: "Lunch out must be after Check-in." Save is blocked. Offending field is highlighted', "High"),
        ("6.8",  "Time correction - missing prereq","Enter a Phase 4 time without providing Phase 3",
                 'Error: "Return from lunch time is required when recording checkout." Save is blocked', "Med"),
        ("6.9",  "Non-reporting day toggle",       "Toggle a weekday as a no-reporting day from the Roster tab",
                 "Day is added to the no_report_days table. Calendar, attendance rate calculations, and exports exclude this day", "High"),
        ("6.10", "Auto-absent on date change",     "Leave the admin app open past midnight on a reporting day with some reservists unchecked",
                 "At midnight, unchecked reservists (excluding those with approved leaves or pending leave requests) are automatically marked absent for the previous day", "High"),
        ("6.11", "Roster search",                  "Search for a person by partial name in the roster search bar",
                 "List filters in real time. Clear button resets the search", "Low"),
        ("6.12", "Log status filter",              'Filter the daily log by "Present", "MC", and "Absent" status buttons',
                 '"All" resets to unfiltered view. Log shows only matching records for the selected filter', "Low"),
        ("6.13", "Day navigation and cross-cycle", "Use previous/next day arrows to navigate from one cycle's last day to the next cycle's first day",
                 "Active cycle switches automatically. Attendance data loads for the new date. Batch label updates in header", "Med"),
        ("6.14", "Welfare note",                   "Admin writes a welfare note for a specific person on a specific date",
                 'Note is saved to the attendance record. Visible in both the roster and the log view. Toast: "Note saved."', "Med"),
        ("6.15", "Real-time roster update",        "Have a reservist check in while the admin has the overview or log tab open",
                 "Admin's roster updates in real time without a page refresh. The realtime status indicator is green", "High"),
        ("6.16", "Low attendance indicator",       "View a reservist whose attendance rate is below 75% in the roster or log",
                 "Amber attendance percentage and short indicator text appear inline next to the contact number. No separate row or pill is added", "Med"),
        ("6.17", "Late badge with reason",         "View the log entry for a reservist who checked in more than 1 hour late with a written reason",
                 'A late badge appears inline on the log card showing the written reason (e.g. "Late · Dental appointment"). No expansion needed to read it', "Med"),
        ("6.18", "Late badge - no reason given",   "View the log entry for a reservist who was forced through the late reason modal but submitted no reason",
                 'Late badge shows "Late · No reason" inline on the log card', "Med"),
        ("6.19", "Missing clock-out display",      "View a log entry where a reservist is marked Present but has no Phase 4 (end of shift) time",
                 "The OUT column of the log card shows a dash on an orange background. No separate badge row is added below the card", "High"),
        ("6.20", "Mark present voids leave",       "Admin marks a reservist with an approved leave as Present via the roster status button",
                 "Approved leave is voided in the DB. The approvedLeavesCache entry for that person and date is removed immediately. Roster shows Present status", "High"),
    ]),
    ("7. PERSONNEL MANAGEMENT", [
        ("7.1",  "Add new person",                 "Admin adds a new person via the People tab with a valid name, SG contact, and password",
                 'Auth account and personnel record created. Person appears on the roster immediately. Toast: "[Name] added to roster."', "High"),
        ("7.2",  "Add person - auth fails, no orphan","Simulate a DB failure on personnel.add after auth.createUser succeeds (requires mocking)",
                 "Auth account is deleted automatically. Error toast shown. No orphaned auth account remains", "High"),
        ("7.3",  "Add person - duplicate contact", "Attempt to add a person with a contact number already on the current roster",
                 'Error: "This contact is already on the roster." No account or record is created', "High"),
        ("7.4",  "Add person - returning detected","Add a person whose contact matches an existing inactive record",
                 "Re-enrollment prompt appears. Confirming reactivates the old record and assigns it to the current cycle. No new account created", "High"),
        ("7.5",  "Deactivate person",              "Admin removes a reservist from the roster using the deactivate option",
                 "Person is removed from the active roster. is_active set to false in DB. Historical attendance records are preserved", "High"),
        ("7.6",  "Password reset",                 "Admin resets a reservist's password from the roster",
                 "Toast: \"[Name]'s password has been reset.\" Reservist can now log in with the new password", "Med"),
        ("7.7",  "Bulk add personnel",             'Paste a list of "Name, Contact" pairs and confirm bulk add',
                 "Valid entries are added. Invalid lines are skipped. Existing contacts are skipped. Toast summarises added/skipped/failed counts", "Med"),
        ("7.8",  "Person attendance history",      "Open the attendance history modal for a specific person",
                 "Full history loads with dates, statuses, check-in times, and any admin corrections flagged", "Med"),
        ("7.9",  "Collapsible action bar",         "Tap the chevron on a personnel card to expand its action bar, then tap another card's chevron",
                 "Action bar expands on the tapped card. The previously expanded card collapses automatically. Only one card is expanded at a time", "Med"),
        ("7.10", "Avatar lightbox",                "Tap any reservist's profile photo in the roster, overview, or log",
                 "Photo opens in an enlarged lightbox overlay. Tapping outside the photo closes the lightbox", "Low"),
        ("7.11", "Added by / Approved by label",  "View a personnel card for someone added directly by admin vs. someone who self-signed up",
                 '"Added by [Name]" appears for direct adds. "Approved by [Name]" appears for self-signup approvals. Blank if neither applies', "Low"),
        ("7.12", "Click row opens history",        "In the Overview or Log tab, click/tap a person's name or row",
                 "The person's attendance history modal opens directly. No need to navigate to the People tab first", "Med"),
        ("7.13", "Member search - cross-cycle",    "Search for a person who exists in a previous cycle (not the current one)",
                 "Person appears in results with their cycle and status shown. Can be permanently deleted from here", "Med"),
    ]),
    ("8. SUPER ADMIN (MASTER)", [
        ("8.1",  "Add admin account",              "Logged in as Master, add a new admin with a valid name, contact, and password",
                 'Auth account and personnel record (role=admin) created. Appears in admins list. Toast: "[Name] added as admin."', "High"),
        ("8.2",  "Add admin - auth fails, no orphan","Simulate DB failure on personnel.add after auth.createUser succeeds (requires mocking)",
                 "Auth account is deleted automatically. Error toast shown. No orphaned auth account", "High"),
        ("8.3",  "Add admin - access denied",      "Call addAdmin handler while logged in as a regular admin (not Master)",
                 'Error: "Access denied." No account created. Handler returns early even if UI bypass is attempted', "High"),
        ("8.4",  "Promote reservist to admin",     "Logged in as Master, promote an existing reservist to admin",
                 'Person\'s role changes to admin. They disappear from the reservist list and appear in the admins list. Toast: "[Name] promoted to admin."', "High"),
        ("8.5",  "Demote admin to reservist",      "Logged in as Master, remove an admin's supervisor status",
                 "Admin is removed from the admins list and returned to the reservist pool with no batch assignment. Toast confirms", "High"),
        ("8.6",  "Promote/demote - non-master",    "Call confirmPromoteAdmin or confirmDeactivateAdmin while logged in as a regular admin",
                 'Error: "Access denied." No role change occurs', "High"),
        ("8.7",  "Reset supervisor password",      "Logged in as Master, reset a supervisor's password from the Team tab",
                 "Password updated in DB. Supervisor can log in with new password. Toast confirms reset", "Med"),
        ("8.8",  "Department switcher - reload",   "Use the department dropdown in the header to switch from Dept A to Dept B",
                 "All data reloads for Dept B: roster, batches, personnel, pending requests, leave requests. Dept A data is cleared", "High"),
        ("8.9",  "Department switcher - cycle restore","Switch from Dept A to Dept B, navigate to a cycle, then switch back to Dept A",
                 "The last-selected cycle for Dept A is restored when switching back. Each department remembers its own cycle context independently", "Med"),
        ("8.10", "Department switcher - state clear","Enter a search term or apply a filter in Dept A, then switch to Dept B",
                 "Search text and filter state are cleared on switch. No stale text from the previous department appears", "Med"),
        ("8.11", "Cross-dept signup isolation",    "Log in as Master in Dept A. Have a Dept B user submit a signup request",
                 "The Dept B signup does not appear in the Dept A Requests tab. Switching to Dept B shows the signup correctly", "High"),
        ("8.12", "Cross-dept leave isolation",     "Log in as Master in Dept A. Have a Dept B reservist submit a leave request",
                 "The Dept B leave request does not appear in the Dept A Requests tab", "High"),
        ("8.13", "Cross-dept cycle notice isolation","Post a cycle notice in Dept A. Log in as a Dept B reservist",
                 "The Dept A notice banner does not appear on the Dept B reservist's check-in screen", "High"),
        ("8.14", "Cross-dept notification isolation","Enable browser notifications as Master in Dept A. Have a Dept B leave request arrive",
                 "No browser notification fires while the Master is viewing Dept A. Switch to Dept B to see the notification trigger", "Med"),
        ("8.15", "Access: reservist tab block",    "Log in as a reservist and attempt to navigate to the Overview, Roster, Log, or People tab",
                 "Admin tabs are not rendered for reservists. URL-based navigation attempts do not load admin views", "High"),
        ("8.16", "Access: admin RLS - leave",      "Attempt to submit a leave request for a different personnel ID via the API (not through the UI)",
                 "Supabase RLS blocks the insert. Request is rejected. Only the authenticated user can create requests for their own personnel_id", "High"),
        ("8.17", "Access: admin cannot promote",   "Log in as a regular admin and attempt to call the promote or add-admin handler",
                 'Error: "Access denied." No role change occurs regardless of UI state', "High"),
    ]),
    ("9. CYCLE MANAGEMENT", [
        ("9.1",  "Create cycle",                   "Admin creates a new cycle by selecting a start date",
                 "Cycle created with correct label (Cycle N/YYYY), start date (Tuesday), end date (14 days later, Monday), and dekit date (Wednesday). Appears in cycle list", "High"),
        ("9.2",  "Auto cycle creation on login",   "Admin logs in when fewer than 8 future cycles exist",
                 "System automatically creates future cycles to bring the count to 8. No manual action required", "Med"),
        ("9.3",  "Rename cycle",                   "Edit the label of an existing cycle",
                 "Label updates in the header, cycle picker, and all references. DB record updated", "Low"),
        ("9.4",  "Set cycle live",                 "Open the cycle picker and set a different cycle as Live",
                 "Only the selected cycle is now live. The previous live cycle is no longer marked live. Header batch label updates", "High"),
        ("9.5",  "Archive toggle",                 "Toggle the show/hide archived batches option in the cycle picker",
                 "Archived (past) cycles appear or disappear from the list. The toggle state is reflected correctly", "Low"),
        ("9.6",  "Meal allowance toggle",          "Toggle meal allowance on for the current cycle",
                 'Meal timer appears on reservist check-in screen. Meal eligibility badge appears in the admin log after 6h work. Toast: "Meal allowance forms activated."', "Med"),
        ("9.7",  "Cycle notice (broadcast)",       "Post a notice for the current cycle",
                 'Notice appears as a banner on every reservist\'s check-in screen (scoped to that department). Toast: "Notice posted to all reservists."', "Med"),
        ("9.8",  "Bulk no-report days",            "Paste multiple dates in dd/mm/yyyy format and apply them as no-report days",
                 "Valid dates within the cycle are added. Dates outside the cycle range are ignored. Toast reports count added", "Med"),
        ("9.9",  "Jump to date",                   "Use the date jump field to navigate to a specific date in a different cycle",
                 "Active cycle switches to the one containing that date. View offset is set correctly. Attendance for that date loads", "Med"),
    ]),
    ("10. EXPORT AND REPORTS", [
        ("10.1", "Export XLS - data integrity",    "Export attendance for a cycle with known data and open in Excel",
                 "All dates, names, statuses (P/MC/A), totals, and rate percentages are correct. No-report days and public holidays are excluded. Admin-corrected records show P*", "High"),
        ("10.2", "Export XLS - empty cycle",       "Attempt to export a cycle with no reservists",
                 'Error toast: "No reservists found in this cycle." No file is downloaded', "Med"),
        ("10.3", "Export XLS - headers freeze",    "Open the exported Excel file and scroll down past the header row",
                 "The title and column header rows remain visible (frozen panes). Row heights are consistent. Legend and Meal column present", "Med"),
        ("10.4", "Print report preview",           "Open the Print/PDF report for a cycle with attendance data",
                 "Modal opens with formatted report showing all personnel, dates, status codes, summary stats including Meal column. Print button opens browser print dialog", "Med"),
        ("10.5", "Person history export",          "Open a person's history and export it to XLS",
                 "File downloads with name, date, status, in time, out time, and admin-correction flag for each row", "Med"),
        ("10.6", "WhatsApp share",                 "Tap the WhatsApp share button from the overview tab",
                 "Pre-filled attendance summary opens in WhatsApp (or wa.me link). Message contains today's date, present/MC/absent counts, and list of names", "Low"),
    ]),
    ("11. ACCOUNT MANAGEMENT", [
        ("11.1", "Change display name",            "Change the display name from account settings",
                 'Name updates in the header chip and in DB. Toast: "Display name updated."', "Low"),
        ("11.2", "Change password",                "Change password with correct current password and a matching new password (min 6 chars)",
                 'Toast: "Password updated." New password works on next login. Old password no longer works', "High"),
        ("11.3", "Change password - wrong current","Attempt to change password with an incorrect current password",
                 'Error: "Current password is incorrect." Password is not changed', "High"),
        ("11.4", "Upload profile photo",           "Upload a profile photo under 20 MB from account settings",
                 'Avatar updates immediately (optimistic). Uploaded to cloud storage. Toast: "Profile photo updated." Avatar persists on next login', "Low"),
        ("11.5", "Upload photo - over 20 MB",      "Attempt to upload a photo file larger than 20 MB",
                 'Error: "Photo must be under 20 MB." No upload is attempted', "Med"),
        ("11.6", "Tap own photo - lightbox",       "Tap the profile photo shown in account settings or the header chip",
                 "Photo opens in an enlarged lightbox overlay. Tapping outside the photo closes it", "Low"),
        ("11.7", "Remove profile photo",           "Remove the profile photo from account settings",
                 'Avatar is removed from the UI and from cloud storage. Initials placeholder shows instead. Toast: "Profile photo removed."', "Low"),
        ("11.8", "Delete account",                 "Delete account from account settings while online",
                 "Account deactivated in DB. Logged out. Login screen shows deletion confirmation banner. Account no longer accessible", "High"),
        ("11.9", "Delete account - offline",       "Attempt to delete account while the device is offline",
                 'Error: "No connection. Cannot delete account while offline." Account is not deleted', "Med"),
    ]),
    ("12. CALENDAR DISPLAY (RESERVIST)", [
        ("12.1", "Present days",                   "View the attendance calendar for a day where the reservist has a Present status",
                 "Day cell is coloured green", "Med"),
        ("12.2", "MC days",                        "View the attendance calendar for a day with an MC status",
                 "Day cell is amber with a solid border", "Med"),
        ("12.3", "Approved future MC",             "Admin approves an MC request for a future date; reservist views their calendar",
                 "Future date cell is amber with a solid border (same style as a past MC day)", "Med"),
        ("12.4", "Pending future MC",              "Reservist has submitted an MC request that is still pending; views their calendar",
                 "Future date cell has a lighter amber tint with a dashed border, distinguishing it from an approved MC", "Med"),
        ("12.5", "Approved future Personal Leave", "Admin approves a Personal Leave request for a future date; reservist views their calendar",
                 "Future date cell is red with a solid border", "Med"),
        ("12.6", "Pending future Personal Leave",  "Reservist has submitted a Personal Leave request still pending; views their calendar",
                 "Future date cell has a lighter red tint with a dashed border", "Med"),
        ("12.7", "No-report days",                 "View the calendar on a day marked as a no-report day",
                 "Day cell is slate blue", "Low"),
        ("12.8", "Absent days",                    "View the calendar for a day where the reservist was marked absent",
                 "Day cell is red", "Low"),
        ("12.9", "Missing clock-out",              "View the calendar for a day the reservist was present but has no Phase 4 time",
                 "Day cell is orange with a dashed border", "Med"),
        ("12.10","Day detail panel",               "Tap any calendar day (present, absent, MC, leave, no-report, missing clock-out)",
                 "Detail panel appears with correct label, sub-text, and colour matching the day's status. Tapping outside dismisses it", "Med"),
    ]),
    ("13. INFO TAB (RESERVIST)", [
        ("13.1", "Shift info",                     "Log in as a reservist and view the Info tab",
                 "Shift title, phase windows (0900, 1200, 1400, 1800 for Ops), attire, and relevant info items are displayed correctly", "Med"),
        ("13.2", "Meal allowance banner",           "View the Info tab when meal allowance is active vs. when it is off",
                 "Banner shows the active state when meal allowance is on, and an on-hold/inactive state when off", "Med"),
        ("13.3", "Team directory - batchmates only","View the team directory section of the Info tab",
                 "Only other reservists in the same batch are listed. The logged-in user is not shown. Admins are not shown", "Med"),
        ("13.4", "Contact number copy",            "Tap a batchmate's phone number in the team directory",
                 'Number is copied to the clipboard. Toast: "Copied to clipboard."', "Low"),
        ("13.5", "WhatsApp contact link",          "Tap the WhatsApp icon next to a batchmate's number",
                 "WhatsApp opens (or wa.me link is followed) with the correct pre-filled number", "Low"),
        ("13.6", "WA group link conditional",      "View the Info tab when the wa-group-link config is set vs. when it is blank",
                 "WA group link button is shown when configured and hidden when the config value is empty", "Low"),
    ]),
    ("14. PWA AND OFFLINE BEHAVIOUR", [
        ("14.1", "Install to home screen (iOS)",   'On iPhone, tap Share then "Add to Home Screen"',
                 "App icon appears on home screen. Opening it launches in standalone mode (no browser chrome)", "Low"),
        ("14.2", "Install to home screen (Android)",'On Android, tap the install banner or browser menu "Add to Home Screen"',
                 "App icon appears on home screen. Launches in standalone mode", "Low"),
        ("14.3", "Offline page load",              "Open the app with no internet connection (after one prior online load to prime the cache)",
                 "App loads from service worker cache. Offline badge is shown. Last known state is displayed", "High"),
        ("14.4", "A2HS prompt (first visit)",      "Visit the app for the first time on a mobile device that supports A2HS",
                 "Add-to-Home-Screen prompt appears after 30 seconds. Dismissing stores the preference so it does not re-appear within 24 hours", "Low"),
        ("14.5", "Landscape orientation warning",  "Rotate a mobile device to landscape orientation",
                 'Overlay appears: "Please rotate your device to portrait mode." App is unusable in landscape. Rotating back dismisses the overlay', "Low"),
        ("14.6", "In-app browser warning",         "Open the app URL from within WhatsApp or Instagram",
                 "Warning displayed: GPS is blocked inside WhatsApp/Instagram. Instructions shown to open in Safari or Chrome", "Med"),
        ("14.7", "Tab visibility - idle auto-logout","Hide the app tab for 20+ minutes then restore it",
                 "On restore, elapsed time is checked. If over 20 minutes idle, the session is ended and the user is shown the login screen", "High"),
    ]),
]

wb = Workbook()
ws = wb.active
ws.title = "Testing Checklist"

for col, w in zip("ABCDEFG", [4.5, 18.6, 37.0, 32.9, 7.9, 7.9, 21.4]):
    ws.column_dimensions[col].width = w

row = 1

# Title row
ws.row_dimensions[row].height = 24
c = ws.cell(row=row, column=1, value="ReservistGO - Testing Checklist")
c.font = _font(TXT, sz=15, bold=True)
ws.merge_cells(f"A{row}:G{row}")
row += 1

# Thin spacer
ws.row_dimensions[row].height = 4.5
row += 1

# Column headers
ws.row_dimensions[row].height = 16.5
for ci, (h, ah) in enumerate(zip(
        ["ID", "Area", "Test Case", "Expected Result", "Priority", "Result", "Notes / Bug Reference"],
        ["center", "left", "left", "left", "center", "center", "left"]), 1):
    c = ws.cell(row=row, column=ci, value=h)
    c.font = _font(HDR_FG, sz=9, bold=True)
    c.fill = _fill(HDR_BG)
    c.alignment = _al(h=ah, v="center")
    c.border = _br(HDR_BR)
row += 1

for sec_title, data_rows in SECTIONS:
    # Section header
    ws.row_dimensions[row].height = 13.5
    c = ws.cell(row=row, column=1, value=f"  {sec_title}")
    c.font = _font(SEC_FG, sz=10, bold=True)
    c.fill = _fill(SEC_BG)
    c.alignment = _al(h="left", v="center")
    c.border = _sec_br()
    for ci in range(2, 8):
        mc = ws.cell(row=row, column=ci)
        mc.fill = _fill(SEC_BG)
        mc.border = _sec_br()
    ws.merge_cells(f"A{row}:G{row}")
    row += 1

    for i, (id_, area, test_case, expected, priority) in enumerate(data_rows):
        ws.row_dimensions[row].height = 27
        bg = ROW_E if (i % 2 == 1) else ROW_O

        # ID
        c = ws.cell(row=row, column=1, value=id_)
        c.font = _font(TXT_ID, sz=9)
        c.fill = _fill(bg)
        c.alignment = _al(h="center", v="top", wrap=False)
        c.border = _br()

        # Area
        c = ws.cell(row=row, column=2, value=area)
        c.font = _font(TXT, sz=9)
        c.fill = _fill(bg)
        c.alignment = _al(v="top")
        c.border = _br()

        # Test Case
        c = ws.cell(row=row, column=3, value=test_case)
        c.font = _font(TXT, sz=9)
        c.fill = _fill(bg)
        c.alignment = _al(v="top")
        c.border = _br()

        # Expected Result
        c = ws.cell(row=row, column=4, value=expected)
        c.font = _font(TXT, sz=9)
        c.fill = _fill(bg)
        c.alignment = _al(v="top")
        c.border = _br()

        # Priority
        pbg, pfg = {"High": (PH_BG, PH_FG), "Med": (PM_BG, PM_FG), "Low": (PL_BG, PL_FG)}[priority]
        c = ws.cell(row=row, column=5, value=priority)
        c.font = _font(pfg, sz=9, bold=(priority in ("High", "Med")))
        c.fill = _fill(pbg)
        c.alignment = _al(h="center", v="top", wrap=False)
        c.border = _br()

        # Result (blank placeholder)
        c = ws.cell(row=row, column=6, value="")
        c.font = _font("9aa3b2", sz=9)
        c.fill = _fill(ROW_O)
        c.alignment = _al(h="center", v="top")
        c.border = _br()

        # Notes
        c = ws.cell(row=row, column=7, value="")
        c.font = _font(TXT_NT, sz=9)
        c.fill = _fill(bg)
        c.alignment = _al(v="top")
        c.border = _br()

        row += 1

    # Spacer between sections
    ws.row_dimensions[row].height = 6
    row += 1

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "docs", "testing_checklist.xlsx")
wb.save(out)
print(f"Saved: {out}")
